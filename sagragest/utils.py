import openpyxl
import decimal
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from .models import CategoryTemplate, ProductTemplate, CategoryEvent, ProductEvent, Daytime, Order, OrderItem

def import_event_xlsx(file, target_event):
    wb = openpyxl.load_workbook(file)
    errors = []
    created = 0
    updated = 0
    # Mapping logical_id -> oggetto
    ct_map = {}
    pt_map = {}
    ce_map = {}
    pe_map = {}
    dt_map = {}
    o_map = {}
    oi_map = {}
    # 1. CategoryTemplate
    if "CategoryTemplate" in wb.sheetnames:
        ws = wb["CategoryTemplate"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            try:
                description = data.get("description") or ""
                ct, created_ct = CategoryTemplate.objects.update_or_create(
                    name=data["name"],
                    defaults={"description": description}
                )
                ct_map[data["logical_id"]] = ct
                if created_ct:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append(f"CategoryTemplate errore: {data} - {e}")
    # 2. ProductTemplate
    if "ProductTemplate" in wb.sheetnames:
        ws = wb["ProductTemplate"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            try:
                description = data.get("description") or ""
                pt, created_pt = ProductTemplate.objects.update_or_create(
                    name=data["name"],
                    defaults={"description": description}
                )
                pt_map[data["logical_id"]] = pt
                if created_pt:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append(f"ProductTemplate errore: {data} - {e}")
    # 3. CategoryEvent
    if "CategoryEvent" in wb.sheetnames:
        ws = wb["CategoryEvent"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            cat_fk = ct_map.get(data["category"])
            if not cat_fk:
                errors.append(f"CategoryEvent senza categoria valida: {data}")
                continue
            try:
                ce, created_ce = CategoryEvent.objects.update_or_create(
                    event=target_event, category=cat_fk,
                    defaults={
                        "display_order": data.get("display_order"),
                        "display_elements": data.get("display_elements")
                    }
                )
                ce_map[data["logical_id"]] = ce
                if created_ce:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append(f"CategoryEvent errore: {data} - {e}")
    # 4. ProductEvent
    if "ProductEvent" in wb.sheetnames:
        ws = wb["ProductEvent"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            prod_fk = pt_map.get(data["product"])
            cat_fk = ce_map.get(data["category"])
            if not prod_fk or not cat_fk:
                errors.append(f"ProductEvent senza riferimenti validi: {data}")
                continue
            try:
                pe, created_pe = ProductEvent.objects.update_or_create(
                    event=target_event, product=prod_fk, category=cat_fk,
                    defaults={
                        "price": decimal.Decimal(data.get("price") or 0),
                        "display_order": data.get("display_order")
                    }
                )
                pe_map[data["logical_id"]] = pe
                if created_pe:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append(f"ProductEvent errore: {data} - {e}")
    # 5. Daytime
    if "Daytime" in wb.sheetnames:
        ws = wb["Daytime"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            start = data.get("start")
            end = data.get("end")
            note = data.get("note") or ""
            try:
                start_dt = parse_datetime(start) if start else None
                end_dt = parse_datetime(end) if end else None
                if start_dt and timezone.is_naive(start_dt):
                    start_dt = timezone.make_aware(start_dt)
                if end_dt and timezone.is_naive(end_dt):
                    end_dt = timezone.make_aware(end_dt)
                dt, created_dt = Daytime.objects.update_or_create(
                    event=target_event, start=start_dt,
                    defaults={"end": end_dt, "note": note}
                )
                dt_map[data["logical_id"]] = dt
                if created_dt:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append(f"Daytime errore: {data} - {e}")
    # 6. Order
    if "Order" in wb.sheetnames:
        ws = wb["Order"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            daytime_fk = dt_map.get(data["daytime"])
            if not daytime_fk:
                errors.append(f"Order senza Daytime valido: {data}")
                continue
            try:
                number = int(data.get("number") or 0)
            except Exception:
                number = 0
            try:
                created_at = parse_datetime(data.get("created_at")) if data.get("created_at") else None
                closed_at = parse_datetime(data.get("closed_at")) if data.get("closed_at") else None
                if created_at and timezone.is_naive(created_at):
                    created_at = timezone.make_aware(created_at)
                if closed_at and timezone.is_naive(closed_at):
                    closed_at = timezone.make_aware(closed_at)
                notes = data.get("notes") or ""
                client = data.get("client") or ""
                order, created_order = Order.objects.update_or_create(
                    daytime=daytime_fk,
                    event=target_event,
                    number=number,
                    defaults={
                        "created_at": created_at,
                        "closed_at": closed_at,
                        "status": data.get("status", "ORDERED"),
                        "notes": notes,
                        "total": decimal.Decimal(data.get("total") or 0),
                        "table_number": data.get("table_number"),
                        "cover": data.get("cover") or 1,
                        "is_takeaway": bool(data.get("is_takeaway")),
                        "client": client,
                        "extra_price": decimal.Decimal(data.get("extra_price") or 0),
                    },
                )
                o_map[data["logical_id"]] = order
                if created_order:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append(f"Order errore: {data} - {e}")
    # 7. OrderItem
    if "OrderItem" in wb.sheetnames:
        ws = wb["OrderItem"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            order_fk = o_map.get(data["order"])
            prod_event_fk = pe_map.get(data["product_event"])
            if not order_fk or not prod_event_fk:
                errors.append(f"OrderItem senza riferimenti validi: {data}")
                continue
            try:
                quantity = int(data.get("quantity") or 1)
            except Exception:
                quantity = 1
            try:
                price_at_order_time = decimal.Decimal(data.get("price_at_order_time") or 0)
                note = data.get("note") or ""
            except Exception:
                price_at_order_time = decimal.Decimal(0)
                note = ""
            try:
                oi, created_oi = OrderItem.objects.update_or_create(
                    order=order_fk,
                    product_event=prod_event_fk,
                    defaults={
                        "quantity": quantity,
                        "note": note,
                        "price_at_order_time": price_at_order_time,
                    },
                )
                oi_map[data["logical_id"]] = oi
                if created_oi:
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                errors.append(f"OrderItem errore: {data} - {e}")
    return created, updated, errors