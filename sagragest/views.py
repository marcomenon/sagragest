# Standard library
import io
import json
import openpyxl

# Django core
from django.shortcuts import render, get_object_or_404, redirect
from django.utils import timezone
from django.template.loader import render_to_string
from django.http import HttpResponseForbidden, HttpResponse, JsonResponse
from django.views.decorators.http import require_POST, require_GET
from django.contrib import messages
from django.core.management import call_command
from django.db import models
from django import forms

# App imports
from .models import Event, CategoryEvent, CategoryTemplate, ProductTemplate, ProductEvent, Daytime, Order, OrderStatus, OrderItem
from .forms import EventForm, CategoryEventFromTemplateForm, CategoryCreateForm, CategoryEditForm, ProductCreateForm, ProductFromTemplateForm, ProductEditForm
from .resources import EventFullResource
from printers.utils import print_action
from core.utils import authenticated_required, admin_required
from .utils import import_event_xlsx

# Third-party imports
from openpyxl import Workbook
from import_export.formats.base_formats import XLSX as XLSXFormat

# =====================
#   DASHBOARD
# =====================
@authenticated_required
def dashboard(request):
    return render(request, "sagragest/dashboard.html")

# =====================
#   UTILITY FUNCTIONS
# =====================
def get_events_for_user(user):
    if user.is_superuser:
        return Event.objects.all()
    elif user.is_staff:
        return Event.objects.filter(group__in=user.groups.all())
    return Event.objects.none()

def get_event_categories(event):
    return CategoryEvent.objects.filter(event=event).select_related('category').order_by("display_order")

def get_category_products(category):
    return ProductEvent.objects.filter(category=category).select_related("product").order_by("display_order")

# =====================
#   EVENTI
# =====================
@admin_required
def event_detail(request):
    return render(request, "sagragest/events.html")

@admin_required
def event_form(request):
    form = EventForm(user=request.user)
    return render(request, "sagragest/events.html#event-form", {"form": form})

@admin_required
def event_edit(request, pk):
    event = get_object_or_404(get_events_for_user(request.user), pk=pk)
    form = EventForm(instance=event, user=request.user)
    return render(request, "sagragest/events.html#event-form", {"form": form})

@admin_required
def event_table(request):
    events = get_events_for_user(request.user).order_by('-year')
    return render(request, "sagragest/events.html#event-table", {"events": events})

@admin_required
@require_POST
def event_create(request):
    pk = request.POST.get("id")
    if pk:
        event = get_object_or_404(get_events_for_user(request.user), pk=pk)
        form = EventForm(request.POST, instance=event, user=request.user)
    else:
        form = EventForm(request.POST, user=request.user)

    if form.is_valid():
        event = form.save(commit=False)
        if not request.user.is_superuser:
            event.group = request.user.groups.first()
        event.save()
        return event_table(request)
    # Se il form non è valido, mostra il form con gli errori
    return render(request, "sagragest/events.html#event-form", {"form": form})

@admin_required
@require_POST
def event_delete(request, pk):
    event = get_object_or_404(get_events_for_user(request.user), pk=pk)
    event.delete()
    return event_table(request)

@admin_required
def event_import_form(request):
    user = request.user

    if user.is_superuser:
        events = Event.objects.all().order_by("-year")
    else:
        events = Event.objects.filter(group__in=user.groups.all()).order_by("-year")

    return render(request, "sagragest/events.html#event-import-form", {
        "events": events,
    })

@admin_required
@require_POST
def event_import(request):
    source_id = request.POST.get("source_event_id")
    target_id = request.POST.get("target_event_id")

    user = request.user

    if not user.is_superuser:
        allowed_ids = Event.objects.filter(group__in=user.groups.all()).values_list("id", flat=True)
        if int(source_id) not in allowed_ids or int(target_id) not in allowed_ids:
            return HttpResponseForbidden("Non hai i permessi per eseguire questa importazione.")

    call_command("clone_event_data", source_id, target_id)

    # Ritorna la tabella aggiornata dell’evento di destinazione
    event = get_object_or_404(Event, pk=target_id)
    events = get_events_for_user(user)
    return render(request, "sagragest/events.html#event-table", {"events": events})

@admin_required
def event_export_xlsx(request):
    event_id = request.GET.get("event_id")
    if not event_id:
        return HttpResponse("ID evento mancante", status=400)
    event = get_object_or_404(Event, pk=event_id)
    resource = EventFullResource()
    datasets = resource.export_event_full(event)
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, dataset in datasets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        xlsx_format = XLSXFormat()
        tmp_wb = xlsx_format.export_data(dataset)
        # Carica i dati dal file temporaneo
        tmp = openpyxl.load_workbook(io.BytesIO(tmp_wb))
        tmp_ws = tmp.active
        for row in tmp_ws.iter_rows(values_only=True):
            ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=evento_{event.id}_export_full.xlsx'
    return response

class EventImportXLSXForm(forms.Form):
    file = forms.FileField(label="File XLSX", required=True)

@admin_required
def event_import_xlsx_form(request):
    user = request.user
    if user.is_superuser:
        events = Event.objects.all().order_by("-year")
    else:
        events = Event.objects.filter(group__in=user.groups.all()).order_by("-year")
    form = EventImportXLSXForm()
    return render(request, "sagragest/events.html#event-import-xlsx-form", {"form": form, "events": events})

@admin_required
def event_import_xlsx(request):
    if request.method == "POST":
        form = EventImportXLSXForm(request.POST, request.FILES)
        target_event_id = request.POST.get("target_event_id")
        target_event = Event.objects.filter(id=target_event_id).first()
        if not target_event:
            messages.error(request, "Evento di destinazione non trovato")
            return redirect('event-detail')
        if form.is_valid():
            file = form.cleaned_data["file"]
            created, updated, errors = import_event_xlsx(file, target_event)
            msg = f"Importazione completata. Creati: {created}, Aggiornati: {updated}."
            if errors:
                msg += f" Errori: {len(errors)}. {errors[:3]}{'...' if len(errors)>3 else ''}"
                print(f"Errori durante l'importazione: {errors}")
                messages.error(request, msg)
            else:
                messages.success(request, msg)
            return redirect('event-detail')
        else:
            messages.error(request, "File non valido")
            return redirect('event-detail')
    else:
        form = EventImportXLSXForm()
        return render(request, "sagragest/events.html#event-import-xlsx-form", {"form": form, "events": Event.objects.all().order_by('-year')})

@admin_required
def event_export_xlsx_form(request):
    user = request.user
    if user.is_superuser:
        events = Event.objects.all().order_by("-year")
    else:
        events = Event.objects.filter(group__in=user.groups.all()).order_by("-year")
    return render(request, "sagragest/events.html#event-export-xlsx-form", {"events": events})

# =====================
#   CATEGORIE
# =====================
@admin_required
def category_list(request):
    user = request.user
    events = get_events_for_user(user).filter(active=True).order_by("-year")
    event_categories = [
        (event, get_event_categories(event))
        for event in events
    ]
    if request.htmx:
        return render(request, "sagragest/categories.html#category-list", {"event_categories": event_categories})
    return render(request, "sagragest/categories.html", {"event_categories": event_categories})

@admin_required
def category_create_form(request):
    event_id = request.GET.get("event")
    event = get_object_or_404(Event, pk=event_id)
    form = CategoryCreateForm(event=event)
    return render(request, "sagragest/categories.html#category-create-form", {"form": form, "event": event})

@admin_required
@require_POST
def category_create(request):
    event_id = request.GET.get("event")
    event = get_object_or_404(Event, pk=event_id)
    form = CategoryCreateForm(request.POST, event=event)
    if form.is_valid():
        template = form.save()
        CategoryEvent.objects.create(event=event, category=template, display_order=form.cleaned_data.get("display_order", 0))
        categories = get_event_categories(event)
        return render(request, "sagragest/categories.html#category-table", {"event": event, "categories": categories})
    return render(request, "sagragest/categories.html#category-create-form", {"form": form, "event": event})

@admin_required
def category_from_template_form(request):
    event_id = request.GET.get("event")
    event = get_object_or_404(Event, pk=event_id)
    form = CategoryEventFromTemplateForm(event=event)
    return render(request, "sagragest/categories.html#category-from-template-form", {"form": form, "event": event})

@admin_required
@require_POST
def category_from_template_create(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    form = CategoryEventFromTemplateForm(request.POST, event=event)
    if form.is_valid():
        template = form.cleaned_data["category_template"]
        display_order = form.cleaned_data["display_order"]
        if not CategoryEvent.objects.filter(event=event, category=template).exists():
            CategoryEvent.objects.create(event=event, category=template, display_order=display_order or 0)
        categories = get_event_categories(event)
        return render(request, "sagragest/categories.html#category-table", {"event": event, "categories": categories})
    return render(request, "sagragest/categories.html#category-from-template-form", {"form": form, "event": event})

@admin_required
def category_edit_form(request, pk):
    category_event = get_object_or_404(CategoryEvent, pk=pk)
    form = CategoryEditForm(instance=category_event)
    if request.method == "POST":
        form = CategoryEditForm(request.POST, instance=category_event)
        if form.is_valid():
            form.save()
            return render(request, "sagragest/categories.html#category-row", {"category": category_event})
    return render(request, "sagragest/categories.html#category-edit-form", {"category": category_event, "form": form})

@admin_required
@require_POST
def category_delete(request, pk):
    category_event = get_object_or_404(CategoryEvent, pk=pk)
    if not request.user.is_superuser and category_event.event.group not in request.user.groups.all():
        return HttpResponseForbidden()
    category_event.delete()
    return HttpResponse("", content_type="text/html")

@admin_required
@require_POST
def delete_unused_category_templates(request):
    unused_templates = CategoryTemplate.objects.filter(
        ~models.Exists(CategoryEvent.objects.filter(category=models.OuterRef('pk')))
    )
    count = unused_templates.count()
    unused_templates.delete()
    messages.success(request, f"{count} template(s) eliminati con successo.")
    return category_list(request)

# =====================
#   PRODOTTI
# =====================
@admin_required
def product_list(request):
    if request.htmx:
        user = request.user
        if user.is_superuser:
            events = Event.objects.filter(active=True).order_by("-year")
        else:
            events = Event.objects.filter(active=True, group__in=user.groups.all()).order_by("-year")

        event_products = []
        for event in events:
            categories = get_event_categories(event)
            category_products = []
            for category in categories:
                products = get_category_products(category)
                category_products.append((category, products))
            event_products.append((event, category_products))

        return render(request, "sagragest/products.html#product-list", {
            "event_products": event_products
        })

    return render(request, "sagragest/products.html")

@admin_required
def product_list_for_category(request, category_id):
    category = get_object_or_404(CategoryEvent, id=category_id)
    products = get_category_products(category)
    return render(request, "sagragest/products.html#product-table", {
        "category": category,
        "products": products
    })

@admin_required
def product_create_form(request):
    category_id = request.GET.get("category")
    category = get_object_or_404(CategoryEvent, id=category_id)
    form = ProductCreateForm(category_event=category)
    return render(request, "sagragest/products.html#product-create-form", {
        "form": form,
        "category": category
    })

@admin_required
@require_POST
def product_create(request):
    category_id = request.POST.get("category_event")
    category = get_object_or_404(CategoryEvent, id=category_id)
    form = ProductCreateForm(request.POST, category_event=category)

    if form.is_valid():
        product_template = form.save()
        ProductEvent.objects.create(
            event=category.event,
            category=category,
            product=product_template,
            price=form.cleaned_data["price"],
            display_order=form.cleaned_data["display_order"]
        )
    return product_list_for_category(request, category.id)

@admin_required
def product_edit_form(request, pk):
    product_event = get_object_or_404(ProductEvent, pk=pk)
    form = ProductEditForm(instance=product_event)
    return render(request, "sagragest/products.html#product-edit-form", {
        "form": form,
        "product": product_event
    })

@admin_required
@require_POST
def product_edit(request, pk):
    product_event = get_object_or_404(ProductEvent, pk=pk)
    form = ProductEditForm(request.POST, instance=product_event)
    if form.is_valid():
        form.save()
        return render(request, "sagragest/products.html#product-row", {"product": product_event})
    return render(request, "sagragest/products.html#product-edit-form", {
        "form": form,
        "product": product_event
    })

@admin_required
@require_POST
def product_delete(request, pk):
    product_event = get_object_or_404(ProductEvent, pk=pk)
    product_event.delete()
    return HttpResponse("", content_type="text/html")

@admin_required
def product_from_template_form(request):
    category_id = request.GET.get("category")
    category = get_object_or_404(CategoryEvent, pk=category_id)
    form = ProductFromTemplateForm(category_event=category)
    return render(request, "sagragest/products.html#product-from-template-form", {
        "form": form,
        "category": category
    })

@admin_required
@require_POST
def product_from_template_create(request, category_id):
    category = get_object_or_404(CategoryEvent, pk=category_id)
    form = ProductFromTemplateForm(request.POST, category_event=category)

    if form.is_valid():
        template = form.cleaned_data["template"]
        price = form.cleaned_data["price"]
        display_order = form.cleaned_data["display_order"]

        # Verifica se il prodotto esiste già nella categoria
        if not ProductEvent.objects.filter(category=category, product=template).exists():
            ProductEvent.objects.create(
                event=category.event,
                category=category,
                product=template,
                price=price,
                display_order=display_order or 0
            )

    products = get_category_products(category)
    return render(request, "sagragest/products.html#product-table", {
        "category": category,
        "products": products
    })

@admin_required
@require_POST
def delete_unused_product_templates(request):
    unused_templates = ProductTemplate.objects.filter(
        ~models.Exists(ProductEvent.objects.filter(product=models.OuterRef('pk')))
    )
    count = unused_templates.count()
    unused_templates.delete()

    messages.success(request, f"{count} template(s) eliminati con successo.")
    return product_list(request)

# =====================
#   GIORNATE
# =====================
@authenticated_required
def current_daytime(request):
    user = request.user

    if user.is_superuser:
        events = Event.objects.filter(active=True).order_by("-year")
        active_daytimes = Daytime.objects.filter(end__isnull=True)
        last_closed_daytimes = Daytime.objects.filter(end__isnull=False).order_by("-end")
    else:
        events = Event.objects.filter(active=True, group__in=user.groups.all()).order_by("-year")
        active_daytimes = Daytime.objects.filter(end__isnull=True, event__group__in=user.groups.all())
        last_closed_daytimes = Daytime.objects.filter(end__isnull=False, event__group__in=user.groups.all()).order_by("-end")

    cards = []
    for event in events:
        current = next((d for d in active_daytimes if d.event_id == event.id), None)
        last_closed = next((d for d in last_closed_daytimes if d.event_id == event.id), None)

        cards.append({
            "event": event,
            "current_daytime": current,
            "last_closed_daytime": last_closed,
        })

    context = {
        "cards": cards
    }

    if request.htmx:
        return render(request, "sagragest/daytime.html#daytime-list", context)

    return render(request, "sagragest/daytime.html", context)

@authenticated_required
@require_POST
def start_daytime(request):
    event_id = request.POST.get("event_id")
    reuse = request.POST.get("reuse") == "true"
    event = get_object_or_404(Event, pk=event_id)

    if reuse:
        # Riapertura: rimuovi 'end' da ultima giornata chiusa
        last_closed = Daytime.objects.filter(event=event, end__isnull=False).order_by("-end").first()
        if last_closed:
            last_closed.end = None
            last_closed.save()
    else:
        # Nuova giornata: chiudi eventuali aperte e crea nuova
        Daytime.objects.filter(event=event, end__isnull=True).update(end=timezone.now())
        Daytime.objects.create(event=event, start=timezone.now())

    return daytime_card(request, event_id=event.pk)

@authenticated_required
@require_POST
def close_daytime(request, pk):
    daytime = get_object_or_404(Daytime, pk=pk)
    if not daytime.end:
        daytime.end = timezone.now()
        daytime.save()
    return daytime_card(request, event_id=daytime.event.pk)

@authenticated_required
def daytime_history(request):
    user = request.user
    if user.is_superuser:
        history = Daytime.objects.filter(end__isnull=False).select_related("event").order_by("-start")[:50]
    else:
        history = Daytime.objects.filter(
            end__isnull=False,
            event__group__in=user.groups.all()
        ).select_related("event").order_by("-start")[:50]

    return render(request, "sagragest/daytime.html#daytime-history", {
        "history": history
    })

@authenticated_required
def daytime_card(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    user = request.user

    if not user.is_superuser and event.group not in user.groups.all():
        return HttpResponseForbidden("Non autorizzato")

    current = Daytime.objects.filter(event=event, end__isnull=True).first()
    last_closed = Daytime.objects.filter(event=event, end__isnull=False).order_by("-end").first()

    html = render_to_string("sagragest/daytime.html#daytime-card", {
        "card": {
            "event": event,
            "current_daytime": current,
            "last_closed_daytime": last_closed,
        }
    }, request=request)

    response = HttpResponse(html)
    # Triggera evento personalizzato lato client
    response["HX-Trigger"] = "refreshHistory"
    return response

# =====================
#   ORDINI
# =====================
@authenticated_required
def order_dashboard(request):
    user = request.user

    if user.is_superuser:
        events = Event.objects.filter(active=True)
    else:
        events = Event.objects.filter(active=True, group__in=user.groups.all())

    selected_event = events.first()
    daytimes = Daytime.objects.filter(event=selected_event).order_by("-start")

    # Prima attiva di questo evento, altrimenti la più recente
    selected_daytime = daytimes.filter(end__isnull=True).first() or daytimes.first()

    return render(request, "sagragest/orders.html", {
        "events": events,
        "daytimes": daytimes,
        "selected_event": selected_event,
        "selected_daytime": selected_daytime,
    })

def get_status_styles(status):
    styles = {
        "ORDERED": ("bg-base-200", "bolt"),
        "IN_PREPARATION": ("bg-info", "clock"),
        "READY": ("bg-success", "check-circle"),
        "CANCELLED": ("bg-error", "x-circle"),
        "UNKNOWN": ("bg-neutral", "question-mark-circle"),
    }
    return styles.get(status, styles["UNKNOWN"])

@authenticated_required
def order_list(request):
    daytime_id = request.GET.get("daytime_id")
    orders = Order.objects.none()

    if daytime_id:
        orders = (
            Order.objects.filter(daytime_id=daytime_id)
            .select_related("daytime", "event")
            .prefetch_related("items__product_event")
            .order_by("number")
        )

        for order in orders:
            css_class, icon = get_status_styles(order.status)
            order.status_class = css_class
            order.status_icon = icon

    return render(request, "sagragest/orders.html#order-list", {
        "orders": orders
    })


@authenticated_required
def daytime_options(request):
    event_id = request.GET.get("event_id")
    user = request.user

    if not event_id:
        return HttpResponse("")

    if user.is_superuser:
        daytimes = Daytime.objects.filter(event_id=event_id).order_by("-start")
    else:
        daytimes = Daytime.objects.filter(event_id=event_id, event__group__in=user.groups.all()).order_by("-start")

    selected_daytime = daytimes.first()

    orders = Order.objects.filter(daytime=selected_daytime).select_related("daytime", "event").prefetch_related("items__product_event")

    for order in orders:
        css_class, icon = get_status_styles(order.status)
        order.status_class = css_class
        order.status_icon = icon

    return render(request, "sagragest/orders.html#daytime-select-and-orders", {
        "daytimes": daytimes,
        "selected_daytime": selected_daytime,
        "orders": orders,
    })

@authenticated_required
def order_detail(request, pk):
    order = get_object_or_404(Order.objects.prefetch_related("items__product_event__product"), pk=pk)
    return render(request, "sagragest/orders.html#order-detail", {
        "order": order
    })

@authenticated_required
def order_detail_placeholder(request):
    return HttpResponse('<p class="text-sm italic text-center text-gray-500">Seleziona un ordine per visualizzare i dettagli</p>')

@require_POST
@authenticated_required
def handle_order_action(request):
    order_id = request.POST.get("order_id")
    action = request.POST.get("action")
    order = get_object_or_404(Order, pk=order_id)
    print(f"[HTMX] Azione '{action}' ricevuta per Ordine #{order.number}")

    if action in ("PRINT_FOR_USER", "PRINT_FOR_CATEGORIES"):
        print_action(order, action)

    if action in OrderStatus.values:
        if action == OrderStatus.IN_PREPARATION:
            order.closed_at = None
            print(f"[COMMAND] Mostra '{action}' nel monitor Comande per Ordine #{order.number}")
        elif action == OrderStatus.READY:
            order.closed_at = timezone.now()
            print(f"[COMMAND] Mostra '{action}' nel monitor Comande per Ordine #{order.number}")    
        elif action == OrderStatus.CANCELLED:
            order.closed_at = None
            print_command = "PRINT_FOR_ALL"      
        order.status = action
        order.save()
        if action == OrderStatus.CANCELLED:
            print_action(order, print_command)
        css_class, icon = get_status_styles(order.status)
        order.status_class = css_class
        order.status_icon = icon
        response = render(request, "sagragest/orders.html#order-card", {"order": order} )
        response["HX-Trigger"] = json.dumps({"refresh-detail": {"order_id": order.id}})
        return response
    
    if action == "DELETE":
        order.status = OrderStatus.CANCELLED
        order.save()
        print_action(order, "PRINT_FOR_ALL")
        order.delete()
        response = HttpResponse("")
        response["HX-Trigger"] = json.dumps({f"clear-detail-{order_id}": {}})
        return response
    
    css_class, icon = get_status_styles(order.status)
    order.status_class = css_class
    order.status_icon = icon    
    response = render(request, "sagragest/orders.html#order-card", {"order": order} )
    
    return response

@authenticated_required
def order_entry_view(request, order_id=None):
    user = request.user

    # Recupera gli eventi attivi visibili all'utente
    if user.is_superuser:
        events = Event.objects.filter(active=True)
    else:
        events = Event.objects.filter(active=True, group__in=user.groups.all())

    # Se non ci sono eventi attivi visibili, errore
    if not events.exists():
        options = "event"
        return render(request, "sagragest/order_entry_no_daytime.html", {"options": options})

    # Recupera la giornata attiva tra quegli eventi
    active_daytime = Daytime.objects.filter(end__isnull=True, event__in=events).first()
    if not active_daytime:
        options = "daytime"
        return render(request, "sagragest/order_entry_no_daytime.html", {"options": options})

    # Giornate visibili all'utente
    if user.is_superuser:
        daytimes = Daytime.objects.filter(end__isnull=True, event__in=events)
    else:
        daytimes = [active_daytime]

    # Selezione giornata (da GET) o giornata attiva di default
    selected_daytime_id = request.GET.get("daytime_id")
    order = None
    order_items = []
    if selected_daytime_id:
        selected_daytime = get_object_or_404(Daytime, pk=selected_daytime_id)
    elif order_id:
        order = get_object_or_404(Order.objects.prefetch_related("items__product_event__product"), pk=order_id)
        selected_daytime = order.daytime
    else:
        selected_daytime = active_daytime

    # Carica categorie per evento della giornata selezionata
    categories = get_event_categories(selected_daytime.event)

    # Selezione categoria (da GET) o prima disponibile
    selected_category_id = request.GET.get("category_id")
    selected_category = get_object_or_404(CategoryEvent, pk=selected_category_id) if selected_category_id else categories.first()

    # Carica i prodotti della categoria selezionata
    products = get_category_products(selected_category)

    # Se order_id è presente, carica l'ordine e i suoi item (se non già caricato sopra)
    if order_id and order is None:
        order = get_object_or_404(Order.objects.prefetch_related("items__product_event__product"), pk=order_id)
    if order:
        order_items = [
            {
                "id": item.product_event.id,
                "name": item.product_event.product.name,
                "price": float(item.price_at_order_time),
                "qty": item.quantity,
                "note": item.note or ""
            }
            for item in order.items.all()
        ]

    context = {
        "daytimes": daytimes,
        "selected_daytime": selected_daytime,
        "categories": categories,
        "selected_category": selected_category,
        "products": products,
        "order": order,
        "order_items": order_items,
    }

    # Se richiesta HTMX, ritorna il main content completo per aggiornare header, contenuto e footer
    if request.htmx:
        return render(request, "sagragest/order_entry.html#order-main-content", context)

    return render(request, "sagragest/order_entry.html", context)

@require_GET
@authenticated_required
def get_products_by_category(request):
    category_id = request.GET.get("category_id")
    selected_category = get_object_or_404(CategoryEvent, pk=category_id)
    
    products = get_category_products(selected_category)
    return render(request, "sagragest/order_entry.html#order-product-tiles", {"products": products, "selected_category": selected_category})

@require_POST
@authenticated_required
def submit_order(request):
    data = json.loads(request.body)
    user = request.user
    daytime_id = data.get("daytime_id")
    order_id = data.get("order_id")

    if user.is_superuser:
        # Superuser: usa il daytime_id passato
        if not daytime_id:
            return JsonResponse({"error": "Giornata non selezionata."}, status=400)
        daytime = get_object_or_404(Daytime, pk=daytime_id)
    else:
        # Utente normale: recupera la giornata attiva dell'evento attivo del suo gruppo
        group_ids = user.groups.values_list("id", flat=True)
        active_event = Event.objects.filter(active=True, group__id__in=group_ids).first()
        if not active_event:
            return JsonResponse({"error": "Nessun evento attivo per il tuo gruppo."}, status=400)
        daytime = Daytime.objects.filter(event=active_event, end__isnull=True).first()
        if not daytime:
            return JsonResponse({"error": "Nessuna giornata attiva per l'evento."}, status=400)

    def safe_int(val):
        # Se val è una lista o tupla, prendi il primo elemento
        if isinstance(val, (list, tuple)):
            val = val[0] if val else 0
        if val in [None, '', 'null']:
            return 0
        try:
            return int(val)
        except (ValueError, TypeError):
            return 0

    if order_id:
        # MODIFICA ORDINE
        order = get_object_or_404(Order, pk=order_id)
        client = data.get("client")
        if client in [None, '', 'null']:
            client = ""
        table = safe_int(data.get("table"))
        cover = safe_int(data.get("cover"))
        order.client = client
        order.table_number = table
        order.cover = cover
        order.is_takeaway = data.get("is_takeaway", False)
        order.notes = data.get("notes", "")
        order.total = data.get("total", 0.0)
        order.extra_price = data.get("extra_price", 0.0)
        order.daytime = daytime
        order.event = daytime.event
        order.status = OrderStatus.CANCELLED
        order.created_by = user
        should_print = data.get("print")
        if should_print:
            print_action(order, "PRINT_FOR_ALL")
        order.save()
        order.status = OrderStatus.ORDERED
        order.save()
        # Aggiorna gli item: elimina tutti e ricrea
        order.items.all().delete()
        for item in data.get("items", []):
            product_event = get_object_or_404(ProductEvent, pk=item["id"])
            OrderItem.objects.create(
                order=order,
                product_event=product_event,
                quantity=item["qty"],
                note=item.get("note", ""),
                price_at_order_time=product_event.price,
            )
        if should_print:
            print_action(order, "PRINT_FOR_ALL")
        print(f"[BACKEND] Ordine #{{order.number}} MODIFICATO.")
        return JsonResponse({"success": True, "order_id": order.id})
    else:
        # CREAZIONE ORDINE
        client = data.get("client")
        if client in [None, '', 'null']:
            client = ""
        table = safe_int(data.get("table"))
        cover = safe_int(data.get("cover"))
        order = Order.objects.create(
            number=Order.get_next_number_for_daytime(daytime),
            client=client,
            table_number=table,
            cover=cover,
            is_takeaway=data.get("is_takeaway", False),
            notes=data.get("notes", ""),
            status=OrderStatus.ORDERED,
            daytime=daytime,
            event=daytime.event,
            created_by=user,
            total=data.get("total", 0.0),
            extra_price=data.get("extra_price", 0.0),
        )
        for item in data.get("items", []):
            product_event = get_object_or_404(ProductEvent, pk=item["id"])
            OrderItem.objects.create(
                order=order,
                product_event=product_event,
                quantity=item["qty"],
                note=item.get("note", ""),
                price_at_order_time=product_event.price,
            )
        should_print = data.get("print")
        if should_print:
            print_action(order, "PRINT_FOR_ALL")
        print(f"[BACKEND] Ordine #{{order.number}} creato per giornata {{daytime.id}} e evento '{{daytime.event.name}}'.")
        return JsonResponse({"success": True, "order_id": order.id})
